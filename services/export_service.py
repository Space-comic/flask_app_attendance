"""Excel 导出服务：负责将考勤、活动参与、情绪记录转换为工作簿字节流。"""

import io
from datetime import timedelta

import openpyxl
from flask import has_app_context, current_app
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func

from models import db
from models.attendance import AttendanceRecord
from models.activity import Activity, ActivityParticipant
from models.emotion import EmotionRecord


def _to_local(dt):
    """将 UTC naive 时间转换为业务时区时间。"""
    if not dt:
        return None
    if has_app_context():
        offset = int(current_app.config.get('TIMEZONE_OFFSET_HOURS', 8))
        return dt + timedelta(hours=offset)
    return dt


# Excel 表头统一样式。
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(fill_type='solid', fgColor='2563EB')
CENTER = Alignment(horizontal='center', vertical='center')


def _apply_header(ws, headers):
    """
    为工作表第一行写入表头并套用统一样式。

    参数：
        ws: openpyxl 工作表对象。
        headers: 表头文字列表。
    """
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def _autosize(ws):
    """根据单元格内容长度自动调整列宽，提升导出表格可读性。"""
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 4)


def export_activity_participants_excel(activity_id):
    """导出指定活动的参与名单。"""
    activity = Activity.query.get(activity_id)
    if not activity:
        raise ValueError('activity not found')
    rows = ActivityParticipant.query.filter_by(activity_id=activity_id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '活动名单'
    headers = ['序号', '活动ID', '活动名称', '学号', '姓名', '活动时间']
    _apply_header(ws, headers)

    local_time = _to_local(activity.created_at)
    time_str = local_time.strftime('%Y-%m-%d %H:%M:%S') if local_time else ''
    for index, row in enumerate(rows, 1):
        ws.append([index, activity.id, activity.name, row.student_id, row.student_name, time_str])

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_activity_stats_excel():
    """导出活动参与次数统计及活动列表汇总。"""
    rows = (
        db.session.query(
            ActivityParticipant.student_id,
            ActivityParticipant.student_name,
            func.count(ActivityParticipant.id).label('count'),
        )
        .group_by(ActivityParticipant.student_id, ActivityParticipant.student_name)
        .order_by(func.count(ActivityParticipant.id).desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '参与统计'
    _apply_header(ws, ['排名', '学号', '姓名', '参与次数'])
    for index, row in enumerate(rows, 1):
        ws.append([index, row.student_id, row.student_name, int(row.count)])
    _autosize(ws)

    ws2 = wb.create_sheet('活动列表')
    _apply_header(ws2, ['活动ID', '活动名称', '参与人数', '创建时间'])
    activities = Activity.query.order_by(Activity.created_at.desc()).all()
    for activity in activities:
        count = ActivityParticipant.query.filter_by(activity_id=activity.id).count()
        local_time = _to_local(activity.created_at)
        ws2.append([
            activity.id,
            activity.name,
            count,
            local_time.strftime('%Y-%m-%d %H:%M:%S') if local_time else '',
        ])
    _autosize(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_emotion_records_excel(student_id=None, date_from=None, date_to=None, source=None):
    """
    导出情绪识别记录。

    参数：
        student_id: 可选，按学号筛选。
        date_from/date_to: 可选，按时间范围筛选。
        source: 可选，来源限定为 `attendance` 或 `group_photo`。
    """
    emotion_cn = {
        'angry': '愤怒',
        'disgust': '厌恶',
        'fear': '恐惧',
        'happy': '快乐',
        'neutral': '平静',
        'sad': '悲伤',
        'surprise': '惊讶',
    }
    source_cn = {'attendance': '考勤', 'group_photo': '合照'}

    query = EmotionRecord.query
    if student_id:
        query = query.filter_by(student_id=student_id)
    if source:
        query = query.filter_by(source=source)
    if date_from:
        query = query.filter(EmotionRecord.recorded_at >= date_from)
    if date_to:
        query = query.filter(EmotionRecord.recorded_at <= date_to)
    records = query.order_by(EmotionRecord.recorded_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '情绪记录'
    _apply_header(ws, ['学号', '姓名', '识别时间', '情绪', '来源'])

    for row in records:
        local_time = _to_local(row.recorded_at)
        ws.append([
            row.student_id,
            row.student_name,
            local_time.strftime('%Y-%m-%d %H:%M:%S') if local_time else '',
            emotion_cn.get(row.emotion, row.emotion),
            source_cn.get(row.source, row.source),
        ])
    _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_attendance_excel(student_id=None, date_from=None, date_to=None):
    """
    导出考勤记录 Excel。

    参数：
        student_id: 学号筛选。
        date_from/date_to: 日期范围筛选。

    返回：
        io.BytesIO: 可直接传给 `send_file()` 的内存工作簿。
    """
    from models.user import User

    query = AttendanceRecord.query
    if student_id:
        query = query.filter_by(student_id=student_id)
    if date_from:
        query = query.filter(AttendanceRecord.date >= date_from)
    if date_to:
        query = query.filter(AttendanceRecord.date <= date_to)
    records = query.order_by(AttendanceRecord.date.desc(), AttendanceRecord.check_time.desc()).all()

    # 预先加载涉及到的学生班级信息，避免循环内重复查库。
    student_ids = {row.student_id for row in records}
    class_map = {}
    if student_ids:
        for user in User.query.filter(User.id.in_(student_ids)).all():
            class_map[user.id] = user.class_name or ''

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '考勤记录'
    headers = ['学号', '姓名', '班级', '签到日期', '签到时间', '状态', '方式']
    _apply_header(ws, headers)

    status_map = {'present': '已签到', 'absent': '缺勤'}
    method_map = {'face': '人脸识别', 'manual': '手动'}

    for row_idx, row in enumerate(records, 2):
        local_check_time = _to_local(row.check_time)
        ws.cell(row=row_idx, column=1, value=row.student_id)
        ws.cell(row=row_idx, column=2, value=row.student_name)
        ws.cell(row=row_idx, column=3, value=class_map.get(row.student_id, ''))
        ws.cell(row=row_idx, column=4, value=row.date.isoformat() if row.date else '')
        ws.cell(row=row_idx, column=5, value=local_check_time.strftime('%H:%M:%S') if local_check_time else '')
        ws.cell(row=row_idx, column=6, value=status_map.get(row.status, row.status))
        ws.cell(row=row_idx, column=7, value=method_map.get(row.method, row.method))

    _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
